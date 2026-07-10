# -*- coding: utf-8 -*-
"""Master HS concordance workbook — GENERATED FROM concordance.csv (the single
source of truth) via the dashboard's concordance-backed loaders. Category-first:
each category (fixed order) lists every HS-6 under the HS-4 headings it uses;
HS-6 belonging to other categories or unused are struck through. HS-4 may repeat
across categories. Raw materials on a separate tab.

Scope (keep/exclude/uncertain/sub-scale), category assignment, HS-6
descriptions AND per-code justifications all come from concordance.csv; the full
HS-6 nomenclature (for struck sibling codes) comes from hs6_reference.csv; and
2025 values from the parquet. concordance.csv is the only file you edit.

Run from the repo root: ``python build_master.py``."""
from pathlib import Path
import pandas as pd
import streamlit_dashboard as d
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = str(Path(__file__).resolve().parent)
OUT = ROOT + '/HS_master_concordance.xlsx'

# Justifications come from the concordance `note` column (the single source of
# truth), so concordance.csv is the only file you edit when changing scope.
_conc = pd.read_csv(ROOT + '/concordance.csv', dtype=str).fillna('')
JUST = {hf: note for hf, note in zip(_conc['hs_full'], _conc['note']) if note}

# ---- Scope / status straight from concordance.csv (via the dashboard) --------
EXCLUDE_7JUL = d.EXCLUDED_HS10       # concordance status = exclude
UNCERTAIN_7JUL = d.UNCERTAIN_HS10    # concordance status = uncertain
EXCLUDED_HS6 = d.SUBSCALE_HS6        # sub-scale transformer headings (in DROP_HS6)
HS4 = d.HS4_HEADINGS

raw = pd.read_parquet(ROOT + '/cimt_output/cimt_trade_slim.parquet')
raw = raw[raw['flow'] == 'imports'].copy(); raw['year'] = pd.to_numeric(raw['year'], errors='coerce')
cats = d.load_equipment_categories(d.EQUIPMENT_CATEGORIES_FILE, d._path_signature(d.EQUIPMENT_CATEGORIES_FILE))

# Full HS-6 nomenclature per HS-4 (incl. unused siblings) from hs6_reference.csv
ref = d.load_hs6_reference(d._path_signature(d.HS6_REFERENCE_FILE))   # {hs4: {hs6: desc}}
hs6full = {h4: sorted(m) for h4, m in ref.items()}
hs6desc_ref = {h6: desc for m in ref.values() for h6, desc in m.items()}
hs6desc_ca = raw.groupby('hs6')['hs_description'].first().to_dict()
hs6desc_ca.update(d.HS6_DESC_FIX)    # concordance descriptions (incl. 854449 fix)
def hs6d(h6): return hs6desc_ca.get(h6) or hs6desc_ref.get(h6, '')
h10desc = raw.groupby('hs_full')['hs_full_description'].first().to_dict()
v2025 = raw[raw['year'] == 2025].groupby('hs_full')['value_cad'].sum().to_dict()
h10_by_h6 = {}
for hf in sorted([h for h in raw['hs_full'].dropna().unique() if v2025.get(h, 0) > 0]):
    h10_by_h6.setdefault(hf[:6], []).append(hf)

CATNAME = {
 'Large Power Transformer (≥100 MVA)':'Large power transformer (>100 MVA)',
 'Medium / Substation Transformer':'Medium / substation transformer','High-Voltage Switchgear':'HV switchgear',
 'Medium-Voltage Switchgear':'MV switchgear (incl. LV to meter)','Underground / Submarine Cable':'Underground / submarine cable',
 'Static power converters (incl. HVDC)':'Static power converters (incl. HVDC)','Overhead Conductor':'Overhead conductor',
 'Substation reactive-power equipment (shunt reactors, capacitor banks, SVC/STATCOM)':'Reactive-power equipment',
 'Protection & Control panels':'Protection & control','Disconnect Switches (HV/MV)':'Disconnect switches',
 'Raw Materials':'Raw materials (inputs)'}
whole_map = {}; full_map = {}; hs6_owner = {}
for c in cats:
    nm = CATNAME.get(c['name'], c['name'])
    for h6 in c['hs6']: whole_map[h6] = nm; hs6_owner.setdefault(h6, nm)
    for f in c['full']: full_map[f] = nm; hs6_owner.setdefault(f[:6], nm)
used_hs6 = set(whole_map) | {f[:6] for f in full_map}

EQ_ORDER = ['Large power transformer (>100 MVA)','Medium / substation transformer','HV switchgear',
 'MV switchgear (incl. LV to meter)','Underground / submarine cable','Static power converters (incl. HVDC)',
 'Overhead conductor','Reactive-power equipment','Protection & control','Disconnect switches']
RM = 'Raw materials (inputs)'
CATFILL = {
 'Large power transformer (>100 MVA)':'BDD7EE','Medium / substation transformer':'C6E0B4','HV switchgear':'D9C2E9',
 'MV switchgear (incl. LV to meter)':'FCE4D6','Underground / submarine cable':'C9E5E3','Static power converters (incl. HVDC)':'F8CBD8',
 'Overhead conductor':'FFF2CC','Reactive-power equipment':'D6EAF8','Protection & control':'E6D5F0',
 'Disconnect switches':'DBE5F1','Raw materials (inputs)':'EDE1D2'}

def show_h10(cat, h6):
    out = []
    for hf in h10_by_h6.get(h6, []):
        a = full_map.get(hf) or whole_map.get(h6)
        if a == cat or (a is None and hs6_owner.get(h6) == cat):
            out.append(hf)
    return out
def belongs(cat, h6):
    return bool(show_h10(cat, h6)) or whole_map.get(h6) == cat or (h6 in EXCLUDED_HS6 and hs6_owner.get(h6) == cat)
def status_of(hf, h6):
    if h6 in EXCLUDED_HS6: return ('Not used', 'HS-6 excluded — sub-scale (≤16 kVA)')
    if hf in EXCLUDE_7JUL: return ('Excluded', 'Excluded on 7 July 2026')
    if hf in UNCERTAIN_7JUL: return ('Uncertain', 'Flagged 7 July 2026 (uncertain)')
    if full_map.get(hf) or whole_map.get(h6): return ('In scope', '')
    return ('Not used', 'Out of scope (HS-6 carve-out)')

def cat_rows(cat):
    hs4s = sorted({h4 for h4 in HS4 if any(belongs(cat, h6) for h6 in hs6full.get(h4, []))})
    out = []
    for h4 in hs4s:
        for h6 in sorted(hs6full[h4]):
            h10s = show_h10(cat, h6)
            if h10s:
                for hf in sorted(h10s, key=lambda x: -v2025.get(x, 0)):
                    st, note = status_of(hf, h6)
                    just = JUST.get(hf, '') if st in ('Excluded', 'Uncertain') else ''
                    out.append([cat, h4, HS4[h4], h6, hs6d(h6), hf, h10desc.get(hf, ''), v2025.get(hf, 0)/1e6, st, note, just, st == 'Not used'])
            elif whole_map.get(h6) == cat or hs6_owner.get(h6) == cat:
                out.append([cat, h4, HS4[h4], h6, hs6d(h6), '—', '(no 2025 imports recorded)', 0, 'In scope', 'No 2025 import data', '', True])
            else:
                owner = whole_map.get(h6) or hs6_owner.get(h6)
                if h6 in EXCLUDED_HS6:
                    desc = '(excluded — sub-scale transformer ≤16 kVA)'; note = 'HS-6 excluded — sub-scale (≤16 kVA)'
                    just = 'HS-6 excluded at data load — sub-scale (≤16 kVA) transformers, not grid hardware.'
                elif owner:
                    desc = f'(HS-6 belongs to → {owner})'; note = f'Used in: {owner}'; just = ''
                else:
                    desc = '(HS-6 not used — not in concordance)'; note = 'HS-6 not in concordance'; just = ''
                out.append([cat, h4, HS4[h4], h6, hs6d(h6), '—', desc, 0, 'Other category', note, just, True])
    return out

# ================= workbook =================
NAVY='1F3864'; BLUE='2E5496'; RED='F4B6B6'; AMBER='FCE0B0'; GREY='D9D9D9'; GREYTXT='808080'
hfont=Font(bold=True,color='FFFFFF',size=11); title=Font(bold=True,size=15,color=NAVY); sub=Font(italic=True,size=10,color='555555')
thin=Side(style='thin',color='BFBFBF'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical='center'); ctr=Alignment(horizontal='center',vertical='center',wrap_text=True)
strike=Font(strike=True,color=GREYTXT,size=10)
cols=['Equipment category','HS-4','HS-4 heading','HS-6','HS-6 description','HS-10','HS-10 description',
      '2025 imports ($M)','Status','Change log / note','Reason / justification (scope audit)']

def write_sheet(ws, subtitle, rows):
    ws['A1']=subtitle[0]; ws['A1'].font=title
    ws['A2']=subtitle[1]; ws['A2'].font=sub
    hr=4
    for i,h in enumerate(cols,1):
        c=ws.cell(hr,i,h); c.font=hfont; c.fill=PatternFill('solid',fgColor=BLUE); c.alignment=ctr; c.border=border
    r=hr+1; start=r
    for row in rows:
        cat,h4,h4d,h6,h6d_,hf,hfd,val,st,note,just,struck=row
        vals=[cat,h4,h4d,h6,h6d_,hf,hfd,val,st,note,just]
        for i,v in enumerate(vals,1):
            c=ws.cell(r,i,v); c.border=border; c.alignment=wrap
            c.fill=PatternFill('solid',fgColor=CATFILL.get(cat,'F2F2F2')) if i==1 else PatternFill('solid',fgColor='FFFFFF')
            if i==1: c.font=Font(bold=True,size=9)
            if i==8: c.number_format='#,##0.0'; c.alignment=Alignment(horizontal='right',vertical='center')
            if i in (2,4,6): c.alignment=Alignment(horizontal='center',vertical='center')
            if struck and i in (4,5,6,7): c.font=strike
            if i in (6,7,9,10,11):
                if st=='Excluded': c.fill=PatternFill('solid',fgColor=RED)
                elif st=='Uncertain': c.fill=PatternFill('solid',fgColor=AMBER)
                elif st in ('Not used','Other category') and i in (9,10): c.fill=PatternFill('solid',fgColor=GREY)
            if i==9: c.alignment=ctr; c.font=Font(bold=True,size=9,color=(GREYTXT if st in ('Not used','Other category') else '000000'))
        r+=1
    rv={start+i:rows[i] for i in range(len(rows))}
    def merge(colidx,keyfn):
        run=start; last=keyfn(start)
        for rr in range(start+1,r+1):
            key=keyfn(rr) if rr<r else object()
            if key!=last:
                if rr-1>run:
                    for ci in colidx:
                        ws.merge_cells(start_row=run,start_column=ci,end_row=rr-1,end_column=ci)
                        ws.cell(run,ci).alignment=Alignment(horizontal=('center' if ci in (2,4) else 'left'),vertical='center',wrap_text=True)
                run=rr; last=key
    merge([1], lambda rr: rv[rr][0])                                   # category
    merge([2,3], lambda rr: (rv[rr][0], rv[rr][1]))                    # HS4 within category
    merge([4,5], lambda rr: (rv[rr][0], rv[rr][1], rv[rr][3]))         # HS6 within category
    widths=[30,7,26,9,32,12,40,12,13,26,60]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A5'; ws.auto_filter.ref='A4:'+get_column_letter(len(cols))+str(r-1)

wb=Workbook()
ws1=wb.active; ws1.title='Equipment categories'
eq_rows=[]
for cat in EQ_ORDER: eq_rows += cat_rows(cat)
write_sheet(ws1, (
 'Master HS concordance — grid-equipment categories → HS-4 → HS-6 → HS-10 (2025 imports, CAD)',
 'Categories in fixed order. Under each, every HS-6 of its HS-4 headings is listed; HS-6 that belong to another '
 'category or are unused are struck through (see "Change log / note"). An HS-4 may repeat across categories. '
 'Red = excluded 7 July 2026; amber = flagged 7 July (uncertain); grey strikethrough = not used / other category. '
 'Generated from concordance.csv (source of truth).'),
 eq_rows)

ws2=wb.create_sheet('Raw materials (inputs)')
write_sheet(ws2, (
 'Raw materials & inputs — HS concordance (2025 imports, CAD)',
 'Upstream inputs and components (feedstock, winding wire, insulators/bushings, transformer parts), not installed '
 'grid equipment. Same layout: every HS-6 under each HS-4 listed; codes belonging to equipment categories are struck '
 'through with a cross-reference.'),
 cat_rows(RM))

wb.save(OUT)
from collections import Counter
print('wrote', OUT)
print('equipment rows:', len(eq_rows), '| status:', dict(Counter(x[8] for x in eq_rows)))
print('raw-material rows:', len(cat_rows(RM)))
